import os
import openpyxl

# Create a new Excel workbook and sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "HWP Scripts Content"

# Column titles
ws['A1'] = "Script Filename"
ws['B1'] = "Content"
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 100

# Script files range
script_files = [f"hwp-{i:02}.py" for i in range(0, 13)]  # Generating filenames from hwp-00.py to hwp-12.py

# Row counter
row = 2

# Read each script file and write its content to the workbook
for script in script_files:
    # Check if the file exists
    if os.path.exists(script):
        with open(script, 'r', encoding='utf-8') as file:
            content = file.read()  # Read the file content
            # Insert filename and content into the Excel sheet
            ws[f'A{row}'] = script
            ws[f'B{row}'] = content
            row += 1
    else:
        # If file does not exist, note it in the Excel
        ws[f'A{row}'] = script
        ws[f'B{row}'] = "File does not exist."
        row += 1

# Save the workbook
excel_file_path = "/mnt/data/HWP_Scripts_Contents.xlsx"
wb.save(excel_file_path)
excel_file_path
