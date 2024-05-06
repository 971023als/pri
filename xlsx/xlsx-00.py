from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

def create_excel_privacy_policy():
    wb = Workbook()
    ws = wb.active
    ws.title = "개인정보처리방침"

    # 제목 추가
    ws.append(['개인정보처리방침'])
    ws['A1'].font = Font(size=20, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:F1')

    # 개정 이력 추가
    ws.append(['개정 이력'])
    ws.append(['개정번호', '개정일자', '개정자', '개정사유', '개정내용', '비고'])
    ws.append(['v1.0.0', '00.00.00', '신규', '신규 개정', '-', '-'])

    # 목차 추가
    ws.append(['목차'])
    contents = [
        "제1조(개인정보의 처리목적) 페이지 2",
        "제2조(개인정보의 처리 및 보유기간) 페이지 3",
        "제3조(처리하는 개인정보의 항목) 페이지 5"
    ]
    for content in contents:
        ws.append([content])

    # 개인정보 처리방침 본문
    ws.append(['개인정보처리방침 본문'])
    privacy_policy_text = (
        "는 정보주체의 자유와 권리 보호를 위해 「개인정보 보호법」 및 관계 법령이 정한 바를 준수하여, "
        "적법하게 개인정보를 처리하고 안전하게 관리하고 있습니다. 이에 「개인정보 보호법」 제30조에 따라 정보주체에게 개인정보 처리에 관한 "
        "절차 및 기준을 안내하고, 이와 관련한 고충을 신속하고 원활하게 처리할 수 있도록 하기 위하여 다음과 같이 개인정보 처리방침을 수립 ‧ 공개합니다."
    )
    ws.append([privacy_policy_text])

    # 열 너비 조정
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # 파일 저장
    wb.save("Privacy_Policy.xlsx")

create_excel_privacy_policy()
