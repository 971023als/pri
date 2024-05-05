from openpyxl import Workbook
from openpyxl.styles import Font

def create_xlsx_privacy_policy_section6():
    # 새 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "개인정보 처리방침"

    # 제목 삽입
    ws['A1'] = "제6조(개인정보처리의 위탁에 관한 사항)"
    ws['A1'].font = Font(bold=True, size=14)

    # 내용 삽입
    paragraphs = [
        "① 밥누리 진흥공단은 원활한 개인정보 업무처리를 위하여 다음과 같이 개인정보 처리업무를 위탁하고 있습니다.",
        "② 밥누리 진흥공단은 위탁계약 체결 시 「개인정보 보호법」 제26조에 따라 위탁업무 수행목적 외 개인정보 처리금지, 기술적・관리적 보호조치, 재위탁 제한, 수탁자에 대한 관리・감독, 손해배상 등 책임에 관한 사항을 계약서 등 문서에 명시하고, 수탁자가 개인정보를 안전하게 처리하는지를 감독하고 있습니다.",
        "③ 「개인정보 보호법」 제26조 제6항에 따라 수탁자가 당사의 개인정보 처리업무를 재 위탁하는 경우 동의를 받고 있습니다.",
        "④ 위탁업무의 내용이나 수탁자가 변경될 경우에는 지체 없이 본 개인정보처리방침을 통하여 공개하도록 하겠습니다."
    ]
    
    row = 2
    for paragraph in paragraphs:
        ws[f'A{row}'] = paragraph
        row += 1

    # 테이블 헤더 추가
    headers = ['개인정보 파일명', '수탁 업체', '수탁 업무 내용', '보유 및 이용기간']
    row += 1
    for col, header in enumerate(headers, start=1):
        ws.cell(row=row, column=col, value=header)

    # 테이블 데이터 추가
    data = [
        ("창업 자금 대출 신청자 정보", "해당 취급 은행", "융자지원의 설정, 유지, 이행, 관리 등", "2023.7.17 ~ 2025.7.17"),
        ("회원가입 정보", "Amazon Web Services, Inc", "Amazon 클라우드 컴퓨팅 환경에 개인정보 보관", "회원 탈퇴 또는 위탁 계약 종료 시")
    ]
    row += 1
    for record in data:
        for col, value in enumerate(record, start=1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    # 열 너비 조정
    column_widths = [25, 25, 30, 30]
    for i, column_width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64+i)].width = column_width

    # 파일 저장
    wb.save("Privacy_Policy_Section6.xlsx")

create_xlsx_privacy_policy_section6()
